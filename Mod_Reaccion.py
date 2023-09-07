# CALCULO MODULO DE REACCIÓN DEL SUELO EN LOSAS
print('\nMODULO DE REACCIÓN VERTICAL DEL SUELO')
print('\nDatos iniciales')
from Clases import *
geom = Geometria()
suelo = Suelo()
vesic = Vesic()
braja = Braja()


def Calcular ():
    # DATOS INICIALES
    suelo.u = float (c1ent01.get())                            
    suelo.Es = float (c1ent02.get())                           
    geom.B = float (c1ent03.get())                             
    geom.L = float (c1ent04.get())                              
    braja.K30 = float (c1ent05.get())                           
    suelo.Tipo = float (c1ent06.get())                                      

    # SOLUCIÓN
    # === Metodo de Vesic, A.B 1961 y Klepikov === #
    geom.A = geom.Area(geom.B, geom.L)                                 
    c2ent07.delete(0, 'end');              
    c2ent07.insert (0, "{:.2f}".format(geom.A))

    geom.w = geom.Coeficiente(geom.B, geom.L)                          
    c2ent08.delete(0, 'end');              
    c2ent08.insert (0, "{:.2f}".format(geom.w))
               
    if (geom.L == geom.B):
        vesic.Kz = vesic.LosaCuadrada(suelo.Es, geom.B, suelo.u)                             
        c2ent09.delete(0, 'end');              
        c2ent09.insert (0, "{:.1f}".format(vesic.Kz))

    else:
        vesic.Kz = vesic.LosaRectang(suelo.Es, geom.w, geom.A, suelo.u)     
        c2ent09.delete(0, 'end');              
        c2ent09.insert (0, "{:.1f}".format(vesic.Kz))   
    
    
    vesic.Kxy = 0.75*vesic.Kz                                               
    c2ent10.delete(0, 'end');              
    c2ent10.insert (0, "{:.1f}".format(vesic.Kxy))
    

    # =====  Metodo Braja M. Das  ===== #
    if (suelo.Tipo == 0):
        braja.Kz = braja.SueloCohesivo(braja.K30, geom.B, geom.L)      
        c3ent11.delete(0, 'end');              
        c3ent11.insert (0, "{:.1f}".format(braja.Kz))

    else:
        braja.Kz = braja.SueloGranular(braja.K30, geom.B, geom.L)      
        c3ent11.delete(0, 'end');              
        c3ent11.insert (0, "{:.1f}".format(braja.Kz))

    braja.Kxy= 0.75*braja.Kz
    c3ent12.delete(0, 'end');              
    c3ent12.insert (0, "{:.1f}".format(braja.Kxy))

    print ("\nCálculos realizados con exito")                                            


def Guardar():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    book = Workbook()
    book = load_workbook('D:\APP_CODICI\Plantilla_ModuloReaccion.xlsx')
    sheet = book.active

    suelo.u = float (c1ent01.get());        sheet['G4'] = "%.2f" %suelo.u                            
    suelo.Es = float (c1ent02.get());       sheet['G5'] = "%.0f" %suelo.Es                           
    geom.B = float (c1ent03.get());         sheet['G6'] = "%.2f" %geom.B                            
    geom.L = float (c1ent04.get());         sheet['G7'] = "%.2f" %geom.L                              
    braja.K30 = float (c1ent05.get());      sheet['G8'] = "%.0f" %braja.K30                          
    suelo.Tipo = float (c1ent06.get())                                      

    # SOLUCIÓN
    # === Metodo de Vesic, A.B 1961 y Klepikov === #
    geom.A = geom.Area(geom.B, geom.L);             sheet['G12'] = "%.2f" %geom.A
    geom.w = geom.Coeficiente(geom.B, geom.L);      sheet['G13'] = "%.2f" %geom.w 

    if (geom.L == geom.B):
        vesic.Kz = vesic.LosaCuadrada(suelo.Es, geom.B, suelo.u);             sheet['G14'] = "%.1f" %vesic.Kz
    else:
        vesic.Kz = vesic.LosaRectang(suelo.Es, geom.w, geom.A, suelo.u);      sheet['G14'] = "%.1f" %vesic.Kz       

    vesic.Kxy = 0.75*vesic.Kz;      sheet['G15'] = "%.1f" %vesic.Kxy   

    # =====  Metodo Braja M. Das  ===== #
    if (suelo.Tipo == 0):
        suelo.Tipo = "Cohesivo";                                        sheet['G9'] = suelo.Tipo
        braja.Kz = braja.SueloCohesivo(braja.K30, geom.B, geom.L);      sheet['G18'] = "%.1f" %braja.Kz
    else:
        suelo.Tipo = "Granular";                                        sheet['G9'] = suelo.Tipo
        braja.Kz = braja.SueloGranular(braja.K30, geom.B, geom.L);      sheet['G18'] = "%.1f" %braja.Kz

    braja.Kxy= 0.75*braja.Kz;      sheet['G19'] = "%.1f" %braja.Kxy

    print ("\nLos datos fueron guardados con exito")
    book.save ('D:\APP_REPORTES\Reporte_ModuloReaccion.xlsx')    


def Borrar():
    c1ent01.delete(0, 'end')
    c1ent02.delete(0, 'end')
    c1ent03.delete(0, 'end')
    c1ent04.delete(0, 'end')
    c1ent05.delete(0, 'end')
    c1ent06.delete(0, 'end')

    c2ent07.delete(0, 'end')
    c2ent08.delete(0, 'end')
    c2ent09.delete(0, 'end')
    c2ent10.delete(0, 'end')

    c3ent11.delete(0, 'end')
    c3ent12.delete(0, 'end')


# ====================================================================================== #
from tkinter import *
vent = Tk()
vent.geometry("620x605")
vent.title(" MODULO DE REACCIÓN DEL SUELO ")
vent.iconbitmap('D:\\BIBLIOTECA PERSONAL\\Programación\\Python\\logo-wat.ico')

# Recuadros interiore
rec1 = LabelFrame(vent, text = ' Datos iniciales. ')
rec1.pack()
rec1.place(x=5, y=5, width=400, height=210)

rec2 = LabelFrame(vent, text = ' Método de Vesic, A.B 1961 y Klepikov ')
rec2.pack()
rec2.place(x=110, y=240, width=400, height=150)

rec3 = LabelFrame(vent, text = ' Método SE_C Seguridad Estructural Cimientos / Braja M. Das ')
rec3.pack()
rec3.place(x=110, y=415, width=400, height=90)

rec5 = LabelFrame(vent, text = ' Valor orientado K30 ')
rec5.pack()
rec5.place(x=415, y=5, width=200, height=210) 

rec6 = LabelFrame(vent, text = ' Opciones ')
rec6.pack()
rec6.place(x=110, y=520, width=400, height=55)

# Contenido recuadro 1

c1tex01 = Label(rec1, text = "µ: Coeficiente de Poison"); c1tex01.pack()
c1tex01.place(x=5, y=10, width=280, height=20)
c1ent01 = Entry(rec1, justify=CENTER);                              c1ent01.place(x=295, y=10, width=90, height=20)

c1tex02 = Label(rec1, text = "Es: Módulo de elasticidad del suelo _ kN/m³"); c1tex02.pack()
c1tex02.place(x=5, y=40, width=280, height=20)
c1ent02 = Entry(rec1, justify=CENTER);                              c1ent02.place(x=295, y=40, width=90, height=20)

c1tex03 = Label(rec1, text = "B: Lado corto de la losa _ m");      c1tex03.pack()
c1tex03.place(x=5, y=70, width=280, height=20)
c1ent03 = Entry(rec1, justify=CENTER);                              c1ent03.place(x=295, y=70, width=90, height=20)

c1tex04 = Label(rec1, text = "L: Lado largo de la losa _ m");      c1tex04.pack()
c1tex04.place(x=5, y=100, width=280, height=20)
c1ent04 = Entry(rec1, justify=CENTER);                              c1ent04.place(x=295, y=100, width=90, height=20)

c1tex05 = Label(rec1, text = "K30: Coeficiente de balasto orientado _ MN/m³");  c1tex05.pack()
c1tex05.place(x=5, y=130, width=280, height=20)
c1ent05 = Entry(rec1, justify=CENTER);                                          c1ent05.place(x=295, y=130, width=90, height=20)

c1tex06 = Label(rec1, text = "Tipo de suelo, 0: Suelo cohesivo  1: Suelo granular"); c1tex06.pack()
c1tex06.place(x=5, y=160, width=280, height=20)
c1ent06 = Entry(rec1, justify=CENTER);                                               c1ent06.place(x=295, y=160, width=90, height=20)

# Contenido recuadro 2

c2tex07 = Label(rec2, text = "A: Área losa de cimentación _ m²");       c2tex07.pack();     c2tex07.place(x=5, y=10, width=280, height=20)
c2ent07 = Entry(rec2, justify=CENTER);                                  c2ent07.place(x=295, y=10, width=90, height=20); c2ent07.config(bg="#ecf0f1")

c2tex08 = Label(rec2, text = "w: Coeficiente en función de la relación L/B");       c2tex08.pack();     c2tex08.place(x=5, y=40, width=280, height=20)
c2ent08 = Entry(rec2, justify=CENTER);                                              c2ent08.place(x=295, y=40, width=90, height=20); c2ent08.config(bg="#ecf0f1")

c2tex09 = Label(rec2, text = "Kz: Módulo de reación vertical del suelo _ kN/m³");   c2tex09.pack();     c2tex09.place(x=5, y=70, width=280, height=20)
c2ent09 = Entry(rec2, justify=CENTER);                                              c2ent09.place(x=295, y=70, width=90, height=20); c2ent09.config(bg="#ecf0f1")

c2tex10 = Label(rec2, text = "Kxy: Módulo de reación horizontal del suelo _ kN/m³");   c2tex10.pack();     c2tex10.place(x=5, y=100, width=280, height=20)
c2ent10 = Entry(rec2, justify=CENTER);                                                 c2ent10.place(x=295, y=100, width=90, height=20); c2ent10.config(bg="#ecf0f1")

# Contenido recuadro 3

c3tex11 = Label(rec3, text = "Kz: Módulo de reación vertical del suelo _ kN/m³");       c3tex11.pack();     c3tex11.place(x=5, y=10, width=280, height=20)
c3ent11 = Entry(rec3, justify=CENTER);                                                  c3ent11.place(x=295, y=10, width=90, height=20); c3ent11.config(bg="#ecf0f1")

c3tex12 = Label(rec3, text = "Kxy: Módulo de reación horizontal del suelo _ kN/m³");    c3tex12.pack();     c3tex12.place(x=5, y=40, width=280, height=20)
c3ent12 = Entry(rec3, justify=CENTER);                                                  c3ent12.place(x=295, y=40, width=90, height=20); c3ent12.config(bg="#ecf0f1")

# Datos cuadro 5 - imagen

from PIL import ImageTk, Image
img = Image.open('D:\APP_CODICI\k30.jpg')
img = img.resize((198,204), Image.ANTIALIAS)

img = ImageTk.PhotoImage(img)
lbl_img = Label(rec5, image=img)
lbl_img.pack()

# Recuadro 6 botones

bot1 = Button(rec6, text = 'Calcular', font='Helvetica 8 bold', command=Calcular); bot1.pack()
bot1.place(x=40, y=5, width=80, height=20)

bot2 = Button(rec6, text = 'Guardar .xls', font='Helvetica 8 bold', command=Guardar); bot2.pack()
bot2.place(x=160, y=5, width=80, height=20)

bot3 = Button(rec6, text = 'Borrar', font='Helvetica 8 bold', command=Borrar); bot3.pack()
bot3.place(x=280, y=5, width=80, height=20)

# Firma
label = Label(vent, text = "wilson.taimalc@gmail.com - 2023", font='Arial 7'); label.pack()
label.place(x=10, y=590, width=600, height=10)
vent.mainloop()      